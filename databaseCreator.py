import sys
import psycopg2
import openpyxl

#USED FOR COLORES PRINTS
class bcolors:
    HEADER = '\033[95m'
    OKBLUE = '\033[94m'
    OKGREEN = '\033[92m'
    WARNING = '\033[93m'
    FAIL = '\033[91m'
    ENDC = '\033[0m'
    BOLD = '\033[1m'
    UNDERLINE = '\033[4m'

################################################################################
#                      DATABASE CONNECTION CONSTANTS                           #
################################################################################
DB_HOST = "localhost"
DB_NAME = "COMERC Database"
DB_USER = "postgres"
DB_PASS = "MyDatabase"

################################################################################
#                         DATABASE CREATION FUNCTIONS                          #
################################################################################
def createRalieTableCommand(name_table):
    command = '''CREATE TABLE ''' + name_table + '''
        (previsao           INT                     NOT NULL,
         ceg                TEXT                    NOT NULL,
         tipo_geracao       TEXT                    NOT NULL,
         combustivel        TEXT                    NOT NULL,
         UF                 TEXT                    NOT NULL,
         usina              TEXT                    NOT NULL,
         potencia_ug        INT                     NOT NULL,
         situacao_das_obras TEXT                    NOT NULL,
         inicio_das_obras   DATE                             ); '''
    return command

def createCadEstadoSubmercadoTableCommand(name_table):
    command = '''CREATE TABLE ''' + name_table + '''
        (estado     TEXT    PRIMARY KEY     NOT NULL,
         sigla      TEXT                    NOT NULL,
         submercado TEXT                    NOT NULL); '''
    return command

#All CREATE TABLE functions must be here for the creation of the database.
#Each element in CREATE_TABLE_COMMANDS is a list containing the name of the 
#table on first position and the method that create its respective CREATE TABLE
#command. 
CREATE_TABLE_COMMANDS = [
    ["ralie", createRalieTableCommand],
    ["cad_estado_submercado", createCadEstadoSubmercadoTableCommand],
]
def createTable(DATABASE_CONNECTION, name_table, command):
    try:
        cur = DATABASE_CONNECTION.cursor()
        cur.execute(command)
        DATABASE_CONNECTION.commit()
    except:
        print(f"{bcolors.WARNING}ALGO DEU ERRADO NA CRIAÇÃO DA TABELA:{bcolors.ENDC}", name_table)
        print("verifique seu banco de dados, se a tabela já existia provavelmente o erro estará relacionado a isso! Em muitos casos não será problema.")
        print()
        DATABASE_CONNECTION.rollback()

################################################################################
#                       DATABASE CONNECTION FUNCTIONS                          #
################################################################################
def connectToDatabase():
    print()
    try:
        DATABASE_CONNECTION = psycopg2.connect(
            host = DB_HOST,
            database = DB_NAME,
            user = DB_USER,
            password = DB_PASS
        )
        print(f"{bcolors.OKGREEN}DATABASE CONNECTED!{bcolors.ENDC}")
        print()
        return DATABASE_CONNECTION
    except:
        print(f"{bcolors.FAIL}DATABASE FAILED TO CONNECT!{bcolors.ENDC}")
        return None

################################################################################
#                         DATABASE INSERT FUNCTIONS                            #
################################################################################
def insertOnTable(DATABASE_CONNECTION, table_name, commands_and_data):
    cur = DATABASE_CONNECTION.cursor()
    i = 0
    n = len(commands_and_data[0])
    while i < n:
        try:
            cur.execute(commands_and_data[0][i], commands_and_data[1][i])
            DATABASE_CONNECTION.commit()
        except psycopg2.Error:
            DATABASE_CONNECTION.rollback()
        i += 1

def genInsertComands(table_name, sheet = None) -> list:
    if table_name == "cad_estado_submercado":
        return genCadEstSubCommands()
    elif table_name == "ralie":
        return genRalieCommands(sheet)
    else:
        print("INSERT GENERATOR FOR TABLE: " + table_name + " DOESN'T EXIST!")

def genCadEstSubCommands() -> list:
    commands = []
    data = []
    cad_estado_file = open("cad_estado_submercado.txt", "r", encoding='UTF-8')
    for line in cad_estado_file:
        #CREATE COMMAND FOR EXECUTE
        splitted_line = line.split(", ")
        commands.append("INSERT INTO cad_estado_submercado (estado, sigla, submercado) VALUES(%s, %s, %s);")
        #CREATE DATA FOR EXECUTE
        data.append((splitted_line[0], splitted_line[1], splitted_line[2].rstrip("\n")))
    return [commands, data]

def genRalieCommands(sheet) -> list:
    commands = []
    data = []
    columns_idx = [1, 3, 4, 5, 6, 7, 10, 13, 19]
    i = 2
    n = sheet.max_row
    while i <= n:
        #CREATE COMMAND FOR EXECUTE
        commands.append("INSERT INTO ralie (previsao, ceg, tipo_geracao, combustivel, UF, usina, potencia_ug, situacao_das_obras, inicio_das_obras) VALUES(%s, %s, %s, %s, %s, %s, %s, %s, %s);")
        #CREATE DATA FOR EXECUTE
        data_list = []
        for j in columns_idx:
            data_list.append(sheet.cell(row=i, column=j).value)
        data.append(tuple(data_list))
        i += 1
    return [commands, data]
################################################################################
#                            XLSX OPERATIONS                                   #
################################################################################
def openRalieXlsx(xlsx_file):
    ralie_xlsx_file = openpyxl.load_workbook(xlsx_file)
    return ralie_xlsx_file

def createHeaders(sheet) -> list:
    headers = []
    j = 1
    headers_num = sheet.max_column
    while j <= headers_num:
        headers.append(sheet.cell(row=1, column=j).value)
        j += 1
    return headers

################################################################################
#                                   MAIN                                       #
################################################################################
def main():
    #PROGRAM CALL CHECK
    if len(sys.argv) < 2:
        print("To create a database of xlsx_file use: >python databaseCreator.py <xlsx_file_path>")
        return
    #DATABASE CONNECTION and TABLE CREATION
    DATABASE_CONNECTION = connectToDatabase()
    for element in CREATE_TABLE_COMMANDS:
        createTable(DATABASE_CONNECTION, element[0], element[1](element[0]))
    #XLSX FILE READ
    xlsx_file = sys.argv[1]
    ralie_xlsx_file = openRalieXlsx(xlsx_file)
    sheet = ralie_xlsx_file["Usinas em Implantação"]
    #XLSX FILE PARSE
    headers = createHeaders(sheet)
    #DATABASE INSERTS
    insertOnTable(DATABASE_CONNECTION, "cad_estado_submercado", genInsertComands("cad_estado_submercado"))
    insertOnTable(DATABASE_CONNECTION, "ralie", genInsertComands("ralie", sheet))

if __name__ == "__main__":
    main()